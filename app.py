"""
Evony Gear Selector - Flask Backend v3
"""
import os
import sys
import json, re, threading, webbrowser, itertools
from collections import Counter, defaultdict
from flask import Flask, render_template, jsonify, request
from openpyxl import load_workbook

# When frozen by PyInstaller, set template/static folders explicitly
# so Flask finds them in _MEIPASS (the extraction folder)
if getattr(sys, 'frozen', False):
    _meipass = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    app = Flask(__name__,
        template_folder = os.path.join(_meipass, 'templates'),
        static_folder   = os.path.join(_meipass, 'static'))
else:
    app = Flask(__name__)
# Resolve the correct base directory:
# - When frozen by PyInstaller: use the directory of the .exe
# - When running as a script: use the directory of app.py
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE     = os.path.join(BASE_DIR, 'evony_data.xlsx')
SETTINGS_FILE = os.path.join(BASE_DIR, 'settings.json')
LOG_FILE      = os.path.join(BASE_DIR, 'startup.log')

# GitHub update configuration
GITHUB_USER     = 'QAEvony1681'
GITHUB_REPO     = 'EvonyGearOptimizer'
GITHUB_API      = f'https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/releases/latest'
# For private repos, this points to the latest release asset.
# For public repos you can use the raw branch URL instead.
GITHUB_DATA_URL = f'https://github.com/{GITHUB_USER}/{GITHUB_REPO}/releases/latest/download/evony_data.xlsx'
APP_VERSION     = '1.3.0'

# In-memory data cache
_data_cache = None

def log(msg):
    """Append a timestamped message to startup.log for troubleshooting."""
    import datetime
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")
    except Exception:
        pass   # logging must never crash the app


def get_cached_data():
    global _data_cache
    if _data_cache is None:
        _data_cache = load_all_data_from_disk()
    return _data_cache

def invalidate_cache():
    global _data_cache
    _data_cache = None
SLOTS = ['Helmet', 'Armor', 'Leg Armor', 'Boots', 'Ring', 'Weapon']
TOP_N_PER_SLOT = 6

# Priority weight by position in Mapping column (1-indexed)
def priority_weight(position):
    if position <= 2: return 3.0
    if position <= 4: return 2.0
    if position <= 6: return 1.5
    return 1.0

# ── Data helpers ──────────────────────────────────────────────────
def parse_buff_value(text):
    if not text: return 0.0
    m = re.search(r'([+-]?\d+(?:\.\d+)?)', str(text))
    return float(m.group(1)) if m else 0.0

def normalize_buff_name(text):
    if not text: return ''
    return re.sub(r'\s*[+-]\d+(?:\.\d+)?%?\s*$', '', str(text).strip()).strip()

# ── Buff expansion: split combined buff names into individual entries ──────────
#
# Handles patterns like:
#   "Ground Troop and Mounted Troop Attack"                    -> 2
#   "Ground and Mounted Troop Attack and Defense"              -> 4
#   "Ground and Mounted Troop Attack, Defense and HP"          -> 6
#   "All Troops Attack"                                        -> 4
#   "All Troops Attack and Defense w/ Dragon"                  -> 8
#   "Attacking Troops Attack"                                  -> 4
#   "Enemy Troop Attack"                                       -> 4
#   "Troop Attack, Defense and HP"                             -> 12

# The four individual troop types "All Troops" expands to
_ALL_TROOP_TYPES = ['Ground Troop', 'Mounted Troop', 'Ranged Troop', 'Siege Machine']

# Canonical single forms of named troop/unit tokens
_TROOP_SINGLES = [
    'Ground Troop', 'Mounted Troop', 'Ranged Troop', 'Siege Machine',
    'All Troops', 'Troop', 'Troops', 'Monsters',
]
# Map plural/variant forms -> canonical singular
_TROOP_CANON = {
    'Ground Troops':   'Ground Troop',
    'Mounted Troops':  'Mounted Troop',
    'Ranged Troops':   'Ranged Troop',
    'Siege Machines':  'Siege Machine',
    'Siege machine':   'Siege Machine',
    'All Troop':       'All Troops',
    'Attacking Troops':'Attacking Troop',  # treated as generic below
    'Attacking Troop': 'Attacking Troop',
    'Enemy Troops':    'Enemy Troop',
    'Enemy Troop':     'Enemy Troop',
    'Troop':           'Troop',
    'Troops':          'Troop',
}
# Tokens that mean "expand to all four troop types"
_GENERIC_TROOP_TOKENS = {
    'All Troops', 'Troop', 'Troops',
    'Attacking Troop', 'Attacking Troops',
    'Enemy Troop', 'Enemy Troops',
}
# Stats that can be split
_STATS = {'Attack', 'Defense', 'HP', 'Load'}
# Stats that should NOT expand generically (only meaningful as a single buff)
_NO_EXPAND_STATS = {'Capacity', 'Speed'}  # Load expands to individual troop types
# Suffix phrases that follow the stat (kept verbatim on each split entry)
_STAT_SUFFIXES = [
    'on Monsters', 'w/ Dragon', 'w/Dragon',
    'when Defending', 'When Defending',
]

def _canon_troop(t):
    return _TROOP_CANON.get(t, t)

def _is_generic_token(t):
    """True if this troop token should expand to all four individual types."""
    return t in _GENERIC_TROOP_TOKENS or _TROOP_CANON.get(t, t) in _GENERIC_TROOP_TOKENS

def _split_troop_types(segment):
    """
    Parse a troop-type segment into a list of canonical troop type strings.
    Returns None if not recognisable.
    """
    seg = segment.strip()
    # Direct match against known tokens
    for tt in sorted(list(_TROOP_SINGLES) + list(_TROOP_CANON.keys()), key=len, reverse=True):
        if seg == tt:
            return [_canon_troop(tt)]

    # "[TypeA] and [TypeB] ..."  — try all " and " split points
    if ' and ' in seg:
        idx = seg.find(' and ')
        left  = seg[:idx].strip()
        right = seg[idx+5:].strip()

        # right is a full known type
        for tt in sorted(list(_TROOP_SINGLES) + list(_TROOP_CANON.keys()), key=len, reverse=True):
            if right == tt:
                left_types = _split_troop_types(left)
                if left_types:
                    return left_types + [_canon_troop(tt)]
                break

        # right shares a suffix with left (e.g. "Ground and Mounted Troop")
        right_words = right.split()
        for suffix_len in range(len(right_words), 0, -1):
            suffix        = ' '.join(right_words[-suffix_len:])
            left_full     = (left + ' ' + suffix).strip()
            right_full    = right
            lc = _canon_troop(left_full)
            rc = _canon_troop(right_full)
            left_known  = lc in _TROOP_SINGLES or left_full in _TROOP_CANON
            right_known = rc in _TROOP_SINGLES or right_full in _TROOP_CANON
            if left_known and right_known:
                return [lc, rc]

    return None


def _split_stats(segment):
    """
    Parse a stat segment like "Attack", "Attack and Defense",
    "Attack, Defense and HP" into a list of stat strings.
    Returns None if not recognisable as pure stat words.
    """
    seg = segment.strip()
    if seg in _STATS:
        return [seg]
    parts = re.split(r',\s*|\s+and\s+', seg)
    stats = [p.strip() for p in parts]
    if all(s in _STATS for s in stats):
        return stats
    return None


def _expand_generic(leading, troop_token, stats, trailing):
    """Expand a generic troop token (All Troops / Troop / Enemy Troop etc.)
    to all four individual troop types."""
    # Determine the enemy/attacking prefix to carry through
    prefix = ''
    if 'Enemy' in leading or 'Enemy' in troop_token:
        prefix = 'Enemy '
    elif 'Attacking' in leading or 'Attacking' in troop_token:
        prefix = 'Attacking '
    elif 'In-City' in leading:
        prefix = 'In-City '

    results = []
    for tt in _ALL_TROOP_TYPES:
        for st in stats:
            results.append(f"{prefix}{tt} {st}{trailing}")
    return results


def expand_plural_buff(name):
    """
    Split a combined buff name into individual buff names.
    Returns a list of one or more buff name strings.
    Falls back to [name] if the pattern is not recognised.
    """
    work = name.strip()

    # Strip leading prefix
    leading = ''
    for q in ('WHEN DEFENDING, ', 'In-City ', 'Attacking ', 'Enemy '):
        if work.startswith(q):
            leading = q
            work = work[len(q):]
            break

    # Normalise w/Dragon -> w/ Dragon
    work = work.replace('w/Dragon', 'w/ Dragon')

    # Strip trailing suffix
    trailing = ''
    for suf in _STAT_SUFFIXES:
        if work.endswith(' ' + suf):
            trailing = ' ' + suf
            work = work[:-len(trailing)].strip()
            break

    # Must contain "and" or "," to potentially be combined
    if ' and ' not in work and ',' not in work:
        # Could still be a generic troop token with single stat
        # e.g. "All Troops Attack" or "Troops Defense on Monsters"
        words = work.split()
        for split_pos in range(1, len(words)):
            troop_seg = ' '.join(words[:split_pos])
            stat_seg  = ' '.join(words[split_pos:])
            if stat_seg in _STATS and stat_seg not in _NO_EXPAND_STATS:
                tt = _split_troop_types(troop_seg)
                if tt and len(tt) == 1 and _is_generic_token(tt[0]):
                    return _expand_generic(leading, tt[0], [stat_seg], trailing)
        return [name]

    # Try each word position as troop/stat boundary
    words = work.split()
    troop_types = None
    stats = None

    for split_pos in range(1, len(words)):
        troop_seg = ' '.join(words[:split_pos])
        stat_seg  = ' '.join(words[split_pos:])
        tt = _split_troop_types(troop_seg)
        st = _split_stats(stat_seg)
        if tt and st:
            # Skip if stat is a non-expandable type (Load, Speed, etc.)
            if any(s in _NO_EXPAND_STATS for s in st):
                return [name]
            troop_types = tt
            stats = st
            break

    if not troop_types or not stats:
        return [name]

    # Check if any troop type is generic -> expand to all four
    expanded_types = []
    for tt in troop_types:
        if _is_generic_token(tt):
            expanded_types.extend(_ALL_TROOP_TYPES)
        else:
            expanded_types.append(tt)

    # Cross-product: (troop_type, stat)
    results = []
    for tt in expanded_types:
        for st in stats:
            # Carry the leading prefix only if not already encoded in tt
            pfx = leading if not any(
                x in tt for x in ('Ground', 'Mounted', 'Ranged', 'Siege')
            ) else leading
            results.append(f"{pfx}{tt} {st}{trailing}")

    return results if len(results) > 1 else [name]
def get_buff_entries(raw_list):
    results = []
    for raw in raw_list:
        if not raw or not str(raw).strip(): continue
        raw = str(raw).strip()
        is_flat = bool(re.search(r'[+-]\d+(?:\.\d+)?$', raw) and
                       not re.search(r'[+-]\d+(?:\.\d+)?%', raw))
        value = parse_buff_value(raw)
        name  = normalize_buff_name(raw)
        for exp in expand_plural_buff(name):
            results.append({'name': exp, 'value': value, 'raw': raw, 'is_flat': is_flat})
    return results

def parse_set_bonus(text):
    if not text: return []
    results = []
    for line in text.split('\n'):
        m = re.match(r'^(\d+)\s+pcs?:\s*(.+)$', line.strip(), re.IGNORECASE)
        if not m: continue
        n_pcs = int(m.group(1)); bt = m.group(2).strip()
        vm = re.search(r'([+-]?\d+(?:\.\d+)?)', bt)
        val = float(vm.group(1)) if vm else 0.0   # preserve sign: -10 stays -10
        is_flat = bool(vm) and '%' not in bt
        name = re.sub(r'\s*[+-]?\d+(?:\.\d+)?%?\s*$', '', bt).strip()
        # Expand combined buff names (e.g. "Ground and Mounted Troop Attack and Defense")
        for expanded_name in expand_plural_buff(name):
            results.append({'min_pcs': n_pcs, 'name': expanded_name, 'value': val,
                            'is_flat': is_flat, 'raw': bt})
    return results

def load_mapping(wb):
    """
    Returns {scenario_key: [(buff_name, weight), ...]}
    Reads the rebuilt Mapping sheet where each scenario column is paired with
    a weight column (header contains ★). Row 1 = title, row 2 = headers, row 3+ = data.
    If a weight cell is blank, falls back to priority_weight(position) based on row order.
    """
    ws = wb['Mapping']
    # Row 2 contains the actual column headers (row 1 is the title bar)
    headers = [cell.value for cell in ws[2]]

    # Map scenario names to (buff_col_idx, weight_col_idx), both 0-indexed
    scenario_cols = {}
    i = 1  # skip col 0 (Master List)
    while i < len(headers):
        h = headers[i]
        if not h:
            i += 1; continue
        if '★' in str(h):   # ★ = weight column, skip
            i += 1; continue
        scenario_name = str(h).strip()
        # Check if next column is the paired weight column
        weight_col_idx = None
        if i + 1 < len(headers) and headers[i+1] and '★' in str(headers[i+1]):
            weight_col_idx = i + 1
        scenario_cols[scenario_name] = (i, weight_col_idx)
        i += 2 if weight_col_idx is not None else 1

    mapping = {}
    for scenario_name, (buff_ci, wt_ci) in scenario_cols.items():
        entries = []
        position = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            buff_val = row[buff_ci] if buff_ci < len(row) else None
            if not buff_val:
                continue
            buff_name = str(buff_val).strip()
            # Read explicit weight; fall back to position-based default
            wt_val = None
            if wt_ci is not None and wt_ci < len(row):
                wt_val = row[wt_ci]
            try:
                weight = float(wt_val) if wt_val is not None and str(wt_val).strip() != '' else None
            except (ValueError, TypeError):
                weight = None
            if weight is None:
                weight = priority_weight(position + 1)   # fallback: position-based
            entries.append((buff_name, weight))
            position += 1
        mapping[scenario_name] = entries
    return mapping

def load_menus(wb):
    ws = wb['Menus']
    menus = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        t1 = str(row[0]).strip()
        t2 = str(row[1]).strip().lstrip('→ ') if row[1] else ''
        t3 = str(row[2]).strip().lstrip('→ ') if row[2] else ''
        menus.append({'tier1': t1, 'tier2': t2, 'tier3': t3})
    return menus

def build_piece(source, tier, name, slot, buffs_raw, debuffs_raw,
                set_bonus_text, full_set_own, notes, upgrade, forge_level, verified):
    sb = str(set_bonus_text).strip() if set_bonus_text else ''
    return {
        'source': source, 'tier': tier, 'name': name, 'slot': slot,
        'buffs':   get_buff_entries(buffs_raw),
        'debuffs': get_buff_entries(debuffs_raw),
        'set_bonus': sb, 'set_bonus_parsed': parse_set_bonus(sb),
        'full_set_own': str(full_set_own).strip() if full_set_own else '',
        'notes': str(notes).strip() if notes else '',
        'upgrade': str(upgrade).strip() if upgrade else '',
        'forge_level': forge_level, 'verified': str(verified).strip() if verified else 'No',
        'raw_buffs':  [r for r in buffs_raw if r],
        'raw_debuffs':[r for r in debuffs_raw if r],
    }

def load_forge_gear(wb):
    ws = wb['Forge Gear']
    pieces = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        try: fl = int(row[13]) if row[13] else 0
        except: fl = 0
        pieces.append(build_piece('forge', str(row[0]).strip(), str(row[1]).strip(),
            str(row[2]).strip(), [row[3],row[4],row[5],row[6]], [row[7],row[8],row[9]],
            row[10], '', row[11], row[12], fl, row[14]))
    return pieces

def load_civ_gear(wb):
    ws = wb['Civilization Gear']
    pieces = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        # Column layout after Category insertion (col 1):
        # 0=Tier, 1=Category, 2=Name, 3=Slot, 4-7=Buffs, 8-10=Debuffs,
        # 11=SetBonus, 12=FullSetOwn, 13=Notes, 14=Upgrade, 15=ForgeLevel, 16=Verified
        try: fl = int(row[15]) if row[15] else 0
        except: fl = 0
        category = str(row[1]).strip() if row[1] else ''
        p = build_piece('civ', str(row[0]).strip(), str(row[2]).strip(),
            str(row[3]).strip(), [row[4],row[5],row[6],row[7]], [row[8],row[9],row[10]],
            row[11], row[12], row[16], '', fl, row[14])
        p['category'] = category   # 'Conquest' or 'Supremacy'
        pieces.append(p)
    return pieces

def load_all_data_from_disk():
    wb = load_workbook(DATA_FILE, data_only=True)
    return {'forge': load_forge_gear(wb), 'civ': load_civ_gear(wb),
            'mapping': load_mapping(wb), 'menus': load_menus(wb),
            'generals': load_generals(wb)}

# ── Scoring ───────────────────────────────────────────────────────
def score_piece(piece, relevant_weighted):
    """
    relevant_weighted: dict {buff_name_lower: weight}
    Returns (pct_score, flat_score, matched_buffs, matched_debuffs)
    """
    pct = 0.0; flat = 0.0
    mb = []; mdb = []
    for b in piece['buffs']:
        w = relevant_weighted.get(b['name'].lower(), 0)
        if w:
            mb.append({**b, 'weight': w})
            if b['is_flat']: flat += abs(b['value'])
            else:            pct  += abs(b['value']) * w
    for d in piece['debuffs']:
        w = relevant_weighted.get(d['name'].lower(), 0)
        if w:
            mdb.append({**d, 'weight': w})
            if d['is_flat']: flat += abs(d['value'])
            else:            pct  += abs(d['value']) * w
    return round(pct, 1), round(flat, 0), mb, mdb

def score_set_bonuses(combo_pieces, relevant_weighted):
    tier_counts = Counter(p['tier'] for p in combo_pieces)
    pct = 0.0; flat = 0.0; active = []
    for tier, count in tier_counts.items():
        if count < 2: continue
        bonus_list = next((p['set_bonus_parsed'] for p in combo_pieces if p['tier']==tier), [])
        full_own   = next((p.get('full_set_own','') for p in combo_pieces if p['tier']==tier), '')
        for b in bonus_list:
            if b['min_pcs'] <= count:
                w = relevant_weighted.get(b['name'].lower(), 0)
                if w:
                    abs_val = abs(b['value'])   # scoring always uses positive magnitude
                    if b['is_flat']: flat += abs_val
                    else:            pct  += abs_val * w
                    active.append({**b, 'tier': tier, 'count': count})
        if count >= 6 and full_own:
            vm = re.search(r'([+-]?\d+(?:\.\d+)?)', full_own)
            if vm:
                val = abs(float(vm.group(1)))
                is_flat = '%' not in full_own
                name = re.sub(r'\s*[+-]?\d+(?:\.\d+)?%?\s*$', '', full_own).strip()
                w = relevant_weighted.get(name.lower(), 0)
                if w:
                    if is_flat: flat += abs(val)
                    else:       pct  += abs(val) * w
                    active.append({'tier':tier,'count':count,'min_pcs':6,
                                   'name':name,'value':val,'is_flat':is_flat,'raw':full_own})
    return round(pct,1), round(flat,0), active

def find_best_combination(candidates_by_slot, relevant_weighted):
    slot_lists = [candidates_by_slot.get(s, [])[:TOP_N_PER_SLOT] or [None] for s in SLOTS]
    best_total = -1; best_combo = None
    for combo in itertools.product(*slot_lists):
        real = [p for p in combo if p is not None]
        ind_total = sum(p['_score'] for p in real)
        set_pct, _, _ = score_set_bonuses(real, relevant_weighted)
        total = ind_total + set_pct
        if total > best_total:
            best_total = total; best_combo = combo
    real = [p for p in best_combo if p is not None]
    set_pct, set_flat, active = score_set_bonuses(real, relevant_weighted)
    result = [{'slot': s, 'piece': p} for s, p in zip(SLOTS, best_combo)]
    return result, round(best_total, 1), set_pct, set_flat, active

# ── Routes ────────────────────────────────────────────────────────

def build_upgrade_chains(pieces):
    """
    For every forge piece, reconstruct the full upgrade chain
    (root → ... → piece → ... → terminus) from the partial forward paths
    stored in each piece's upgrade field.
    Returns dict: {piece_name: [ordered list of names in chain]}

    Important: each piece's upgrade string lists ALL downstream pieces
    (e.g. Dragon Spear: '→ Ares Spear → Imperial Spear → Asura Spear').
    Only the FIRST step is the direct successor — the rest are further
    descendants. We use only steps[0] to build the predecessor map so
    the chain reflects the true step-by-step progression.
    """
    # Parse forward steps; store only the direct next piece (steps[0])
    direct_next = {}   # piece_name -> immediate successor name
    for p in pieces:
        if p.get('upgrade') and p['upgrade'] not in ('', '—'):
            steps = [s.strip() for s in p['upgrade'].split('→') if s.strip()]
            if steps:
                direct_next[p['name']] = steps[0]   # direct successor only

    # Build reverse lookup: name -> direct predecessor
    predecessor = {}
    for name, nxt in direct_next.items():
        if nxt not in predecessor:
            predecessor[nxt] = name

    # For display, also build the full ordered chain by walking forward
    # from root using direct_next links
    def full_chain(name):
        # Walk back to root
        back = []
        cur = name
        visited = set()
        while cur in predecessor and cur not in visited:
            visited.add(cur)
            cur = predecessor[cur]
            back.append(cur)
        back.reverse()
        # Walk forward from current piece using direct_next links
        fwd = []
        cur = name
        visited2 = set()
        while cur in direct_next and cur not in visited2:
            visited2.add(cur)
            cur = direct_next[cur]
            fwd.append(cur)
        return back + [name] + fwd

    return {p['name']: full_chain(p['name']) for p in pieces}


def load_generals(wb):
    """Load generals list from the Generals sheet.
    Column layout (row 3 headers):
    0=Name, 1=Culture, 2=Rarity, 3=PrimaryRole, 4=SpecialSkill,
    5=BasicBuffs, 6=BasicDebuffs, 7=CovenantGenerals, 8=Status, 9=Notes
    """
    if 'Generals' not in wb.sheetnames:
        return []
    ws = wb['Generals']
    generals = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        generals.append({
            'name':              str(row[0]).strip(),
            'culture':           str(row[1]).strip() if row[1] else '',
            'rarity':            str(row[2]).strip() if row[2] else '',
            'primary_role':      str(row[3]).strip() if row[3] else '',
            'special_skill':     str(row[4]).strip() if row[4] else '',
            'basic_buffs':       str(row[5]).strip() if row[5] else '',
            'basic_debuffs':     str(row[6]).strip() if row[6] else '',
            'covenant_generals': str(row[7]).strip() if row[7] else '',
            'status':            str(row[8]).strip() if row[8] else '',
            'notes':             str(row[9]).strip() if row[9] else '',
        })
    return generals


def aggregate_combo_stats(pieces, active_bonuses, relevant_weighted=None):
    """
    Aggregate all buffs/debuffs across the combo including set bonuses.
    Returns three sorted lists: (buffs, debuffs, flats)
    Each entry: {name, value (raw %), weighted_pts, has_explicit_weight, sources}
    Sorted by weighted_pts descending.
    """
    buff_totals   = defaultdict(lambda: {'value':0.0,'weighted':0.0,'sources':[],'has_wt':False})
    debuff_totals = defaultdict(lambda: {'value':0.0,'weighted':0.0,'sources':[],'has_wt':False})
    flat_totals   = defaultdict(lambda: {'value':0.0,'weighted':0.0,'sources':[],'has_wt':False})

    def is_debuff(name, val=0):
        n = name.lower()
        # Enemy/WHEN DEFENDING prefix = debuff; negative value = debuff (e.g. set bonuses)
        return n.startswith('enemy ') or n.startswith('when defending') or val < 0

    def add_entry(name, val, source, is_flat, weight=None):
        target = flat_totals if is_flat else (debuff_totals if is_debuff(name, val) else buff_totals)
        raw = abs(val)   # store absolute value; sign is captured by which bucket
        target[name]['value']  += raw
        target[name]['sources'].append(source)
        if weight is not None:
            target[name]['weighted']  += raw * weight
            target[name]['has_wt']     = True
        else:
            target[name]['weighted']  += raw   # 1.0 multiplier if no weight info

    for p in pieces:
        label = p['name']
        for b in p['buffs']:
            w = relevant_weighted.get(b['name'].lower()) if relevant_weighted else None
            add_entry(b['name'], b['value'], label, b['is_flat'], w)
        for d in p['debuffs']:
            w = relevant_weighted.get(d['name'].lower()) if relevant_weighted else None
            add_entry(d['name'], d['value'], label, d['is_flat'], w)

    for b in active_bonuses:
        source = f"{b['tier']} {b['min_pcs']}pc set bonus"
        w = relevant_weighted.get(b['name'].lower()) if relevant_weighted else None
        add_entry(b['name'], b['value'], source, b['is_flat'], w)

    def to_list(d):
        rows = []
        for k, v in d.items():
            rows.append({
                'name':   k,
                'value':  round(v['value'], 1),
                'weighted_pts': round(v['weighted'], 1),
                'has_explicit_weight': v['has_wt'],
                'sources': v['sources'],
            })
        # Sort: explicit-weight items first (by pts desc), then N/A items (by raw% desc)
        return sorted(rows,
            key=lambda x: (x['has_explicit_weight'], x['weighted_pts']),
            reverse=True)

    return to_list(buff_totals), to_list(debuff_totals), to_list(flat_totals)

@app.route('/api/data')
def api_data():
    try:
        data = load_all_data()
        seen = set()
        tier1_options = [m['tier1'] for m in data['menus']
                         if m['tier1'] not in seen and not seen.add(m['tier1'])]
        ADV_FORGE_TIERS = {'Apollo', 'Asura'}
        civ_pieces = [{'tier':p['tier'],'name':p['name'],'slot':p['slot'],
                       'category':p.get('category','')} for p in data['civ']]
        adv_forge_pieces = [{'tier':p['tier'],'name':p['name'],'slot':p['slot']}
                            for p in data['forge'] if p['tier'] in ADV_FORGE_TIERS]
        wb = load_workbook(DATA_FILE, data_only=True)
        generals = load_generals(wb)
        return jsonify({'ok':True, 'menus':data['menus'], 'tier1_options':tier1_options,
                        'civ_pieces':civ_pieces, 'adv_forge_pieces':adv_forge_pieces,
                        'generals': generals})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

def load_all_data():
    """Public interface — returns cached data."""
    return get_cached_data()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/recommend', methods=['POST'])
def api_recommend():
    try:
        body = request.get_json()
        tier1       = body.get('tier1','')
        tier2       = body.get('tier2','')
        tier3       = body.get('tier3','')
        forge_level          = int(body.get('forge_level', 50))
        owned_civ            = set(body.get('owned_civ', []))
        owned_adv_forge      = set(body.get('owned_adv_forge', []))  # Apollo/Asura piece names

        data = load_all_data()

        # Tier1 aliases: Menus sheet name -> Mapping sheet prefix
        TIER1_ALIASES = {'PVP Sub-City': 'Sub-City PVP'}
        t1m = TIER1_ALIASES.get(tier1, tier1)

        # Build candidate keys — also strip trailing " Speed" from tier2
        # since some menu labels say "Construction Speed" but mapping says "Construction"
        t2  = tier2 or ''
        t2s = t2.removesuffix(' Speed') if t2.endswith(' Speed') else t2
        candidates_keys = []
        if tier3 and t2:
            candidates_keys.append(f"{t1m} → {t2} → {tier3}")
            candidates_keys.append(f"{t1m} → {t2s} → {tier3}")
            candidates_keys.append(f"{t1m} - {t2} → {tier3}")
            candidates_keys.append(f"{tier1} → {t2} → {tier3}")
        if t2:
            candidates_keys.append(f"{t1m} → {t2}")
            candidates_keys.append(f"{t1m} → {t2s}")
            candidates_keys.append(f"{tier1} → {t2}")
            candidates_keys.append(f"{tier1} → {t2s}")
        candidates_keys += [t1m, tier1]

        scenario_key = None
        for ck in candidates_keys:
            if ck in data['mapping']:
                scenario_key = ck; break
            for k in data['mapping']:
                if k.lower() == ck.lower():
                    scenario_key = k; break
            if scenario_key: break

        if not scenario_key:
            return jsonify({'ok':False,'error':f'Scenario not found. Tried: {candidates_keys}'}), 400

        # Build weighted relevance dict
        relevant_list = data['mapping'][scenario_key]  # [(name, weight), ...]
        relevant_weighted = {name.lower(): weight for name, weight in relevant_list}
        relevant_buffs = [name for name, _ in relevant_list]

        # Filter gear pool
        # Advanced forge tiers (Apollo/Asura) are filtered by individual owned pieces,
        # same as civ gear.  All other forge tiers are filtered by forge level only.
        ADV_FORGE_TIERS = {'Apollo', 'Asura'}

        # Build set of civ pieces already assigned to generals (excluded from recommendations)
        assigned_civ = set(body.get('assigned_civ', []))

        pool = []
        for p in data['forge']:
            if p['tier'] in ADV_FORGE_TIERS:
                if p['name'] in owned_adv_forge:
                    pool.append(p)
            elif p['forge_level'] <= forge_level:
                pool.append(p)
        for p in data['civ']:
            if p['name'] in owned_civ and p['name'] not in assigned_civ:
                pool.append(p)

        # Score individually — Pass 1: all pieces scored
        all_scored = {}   # name -> scored entry
        for p in pool:
            if p['slot'] not in SLOTS: continue
            score, flat, mb, mdb = score_piece(p, relevant_weighted)
            all_scored[p['name']] = {**p, '_score':score, '_flat':flat, '_mb':mb, '_mdb':mdb}

        # Pass 1: collect top-N per slot by individual score
        by_slot = {s: [] for s in SLOTS}
        for entry in all_scored.values():
            if entry['_score'] > 0 or entry['_flat'] > 0:
                by_slot[entry['slot']].append(entry)
        for s in SLOTS:
            by_slot[s].sort(key=lambda x: x['_score'], reverse=True)

        # Pass 2: find tiers represented in top-N; add any 0-score pieces
        # from those same tiers so set bonuses can be evaluated properly
        represented_tiers = set()
        for s in SLOTS:
            for entry in by_slot[s][:TOP_N_PER_SLOT]:
                represented_tiers.add(entry['tier'])

        for entry in all_scored.values():
            if entry['tier'] in represented_tiers:
                slot = entry['slot']
                already_in = any(e['name'] == entry['name'] for e in by_slot[slot])
                if not already_in:
                    by_slot[slot].append(entry)   # score=0 but set-bonus-eligible

        for s in SLOTS:
            by_slot[s].sort(key=lambda x: x['_score'], reverse=True)

        combo_slots, combo_total, set_pct, set_flat, active = \
            find_best_combination(by_slot, relevant_weighted)

        # Serialize combo
        combo_out = []
        for item in combo_slots:
            p = item['piece']
            if not p:
                combo_out.append({'slot': item['slot'], 'piece': None})
            else:
                sc, fl, mb, mdb = score_piece(p, relevant_weighted)
                combo_out.append({'slot': item['slot'], 'piece': {
                    'source':p['source'],'tier':p['tier'],'name':p['name'],'slot':p['slot'],
                    'score':sc,'flat_score':fl,'matched_buffs':mb,'matched_debuffs':mdb,
                    'raw_buffs':p['raw_buffs'],'raw_debuffs':p['raw_debuffs'],
                    'verified':p['verified'],'forge_level':p['forge_level'],
                    'set_bonus':p['set_bonus'],'upgrade':p.get('upgrade',''),
                    'full_set_own':p.get('full_set_own',''),
                }})

        ranked = {}
        for s in SLOTS:
            ranked[s] = [{'source':p['source'],'tier':p['tier'],'name':p['name'],'slot':p['slot'],
                'score':p['_score'],'flat_score':p['_flat'],'matched_buffs':p['_mb'],
                'matched_debuffs':p['_mdb'],'raw_buffs':p['raw_buffs'],'raw_debuffs':p['raw_debuffs'],
                'verified':p['verified'],'forge_level':p['forge_level'],
                'set_bonus':p['set_bonus'],'upgrade':p.get('upgrade',''),
                'full_set_own':p.get('full_set_own',''),
            } for p in by_slot[s]]

        # Aggregate collective stats for the recommended combo
        combo_real_pieces = [item['piece'] for item in combo_slots if item['piece'] is not None]
        combo_buffs, combo_debuffs, combo_flats = aggregate_combo_stats(combo_real_pieces, active, relevant_weighted)

        return jsonify({'ok':True,'scenario':scenario_key,'relevant_buffs':relevant_buffs,
            'combo':combo_out,'combo_total':combo_total,'combo_set_pct':set_pct,
            'combo_set_flat':set_flat,'active_bonuses':active,'by_slot':ranked,'slots':SLOTS,
            'combo_buffs':combo_buffs,'combo_debuffs':combo_debuffs,'combo_flats':combo_flats})

    except Exception as e:
        import traceback
        return jsonify({'ok':False,'error':str(e),'trace':traceback.format_exc()}), 500

@app.route('/api/piece_detail/<path:piece_name>')
def api_piece_detail(piece_name):
    try:
        data = load_all_data()
        chains = build_upgrade_chains(data['forge'])
        for p in data['forge'] + data['civ']:
            if p['name'] == piece_name:
                full_chain = chains.get(piece_name, [piece_name])
                return jsonify({'ok':True,'piece':p,'upgrade_chain':full_chain})
        return jsonify({'ok':False,'error':'Piece not found'}), 404
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

@app.route('/api/score_set', methods=['POST'])
def api_score_set():
    """
    Score an arbitrary set of up to 6 pieces for a given scenario.
    Body: { scenario_key, piece_names: [str, ...] }
    Returns same shape as the combo section of /api/recommend.
    """
    try:
        body         = request.get_json()
        scenario_key = body.get('scenario_key', '')
        piece_names  = body.get('piece_names', [])   # may include None/empty for empty slots
        forge_level  = int(body.get('forge_level', 50))
        owned_civ    = set(body.get('owned_civ', []))
        owned_adv    = set(body.get('owned_adv_forge', []))

        data = load_all_data()

        if scenario_key not in data['mapping']:
            return jsonify({'ok': False, 'error': f'Scenario not found: {scenario_key}'}), 400

        relevant_weighted = {n.lower(): w for n, w in data['mapping'][scenario_key]}

        # Build piece lookup filtered by ownership/forge level (same rules as recommend)
        ADV_FORGE_TIERS = {'Apollo', 'Asura'}
        available = {}
        for p in data['forge']:
            if p['tier'] in ADV_FORGE_TIERS:
                if p['name'] in owned_adv: available[p['name']] = p
            elif p['forge_level'] <= forge_level:
                available[p['name']] = p
        for p in data['civ']:
            if p['name'] in owned_civ: available[p['name']] = p

        # Resolve selected pieces (skip empty/None entries)
        selected = []
        slot_used = set()
        for name in piece_names:
            if not name: continue
            p = available.get(name)
            if p and p['slot'] not in slot_used:
                selected.append(p)
                slot_used.add(p['slot'])

        # Score individual pieces
        scored = []
        for p in selected:
            sc, fl, mb, mdb = score_piece(p, relevant_weighted)
            scored.append({
                'name': p['name'], 'tier': p['tier'], 'slot': p['slot'],
                'source': p['source'], 'score': sc, 'flat_score': fl,
                'matched_buffs': mb, 'matched_debuffs': mdb,
                'forge_level': p['forge_level'], 'verified': p['verified'],
                'raw_buffs': p['raw_buffs'], 'raw_debuffs': p['raw_debuffs'],
                'set_bonus': p['set_bonus'], 'upgrade': p.get('upgrade', ''),
                'full_set_own': p.get('full_set_own', ''),
            })

        # Set bonuses
        set_pct, set_flat, active = score_set_bonuses(selected, relevant_weighted)
        ind_total = round(sum(s['score'] for s in scored), 1)
        total     = round(ind_total + set_pct, 1)

        # Aggregate stats
        buffs, debuffs, flats = aggregate_combo_stats(selected, active, relevant_weighted)

        # Build slot map (slot -> piece or None) for the 6-slot display
        slot_map = {s: None for s in SLOTS}
        for s in scored:
            slot_map[s['slot']] = s

        return jsonify({
            'ok': True,
            'total': total, 'ind_total': ind_total, 'set_pct': set_pct,
            'set_flat': set_flat, 'active_bonuses': active,
            'slot_map': slot_map,
            'buffs': buffs, 'debuffs': debuffs, 'flats': flats,
        })

    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e),
                        'trace': traceback.format_exc()}), 500


@app.route('/api/slot_options', methods=['POST'])
def api_slot_options():
    """
    Return all available pieces per slot for the compare dropdowns,
    filtered by forge_level + ownership, grouped by tier.
    Body: { forge_level, owned_civ, owned_adv_forge }
    """
    try:
        body        = request.get_json()
        forge_level = int(body.get('forge_level', 50))
        owned_civ   = set(body.get('owned_civ', []))
        owned_adv   = set(body.get('owned_adv_forge', []))

        data = load_all_data()
        ADV_FORGE_TIERS = {'Apollo', 'Asura'}

        # Collect available pieces grouped by slot then tier
        by_slot = {s: {} for s in SLOTS}  # slot -> {tier -> [name, ...]}

        TIER_ORDER = ["King's", 'Dragon', 'Ares', 'Achaemenidae', 'Parthian',
                      'Imperial', 'Asura', 'Apollo']

        for p in data['forge']:
            if p['tier'] in ADV_FORGE_TIERS:
                if p['name'] not in owned_adv: continue
            elif p['forge_level'] > forge_level:
                continue
            slot = p['slot']
            if slot not in by_slot: continue
            tier = p['tier']
            if tier not in by_slot[slot]: by_slot[slot][tier] = []
            by_slot[slot][tier].append(p['name'])

        for p in data['civ']:
            if p['name'] not in owned_civ: continue
            slot = p['slot']
            if slot not in by_slot: continue
            tier = p['tier']
            if tier not in by_slot[slot]: by_slot[slot][tier] = []
            by_slot[slot][tier].append(p['name'])

        # Convert to ordered list of {tier, names} per slot
        result = {}
        for slot, tiers in by_slot.items():
            # Forge tiers in fixed order, then civ tiers alphabetically
            forge_tiers = [(t, tiers[t]) for t in TIER_ORDER if t in tiers]
            civ_tiers   = sorted([(t, tiers[t]) for t in tiers if t not in TIER_ORDER],
                                  key=lambda x: x[0])
            result[slot] = [{'tier': t, 'names': n} for t, n in forge_tiers + civ_tiers]

        return jsonify({'ok': True, 'by_slot': result, 'slots': SLOTS})

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── Profile-aware settings helpers ────────────────────────────────────────────

DEFAULT_PROFILE = 'Default'
PROFILE_KEYS    = ['civOwned', 'advForgeOwned', 'setCollapsed', 'generalAssignments', 'forgeLevel']
GLOBAL_KEYS     = ['currentTheme', 'lastUpdateCheck', 'startupsSinceCheck']

def load_settings_raw():
    """Load raw settings dict from disk. Returns {} on any error."""
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def migrate_settings(raw):
    """
    Migrate flat (pre-v1.3.0) settings to the profile-aware schema.
    Returns the migrated dict, or the existing one if already migrated.
    """
    if 'profiles' in raw:
        return raw   # already migrated

    # Build a Default profile from the flat keys
    profile_data = {k: raw.get(k, {}) for k in PROFILE_KEYS}
    migrated = {
        'activeProfile':  DEFAULT_PROFILE,
        'profiles':       {DEFAULT_PROFILE: profile_data},
        'profileOrder':   [DEFAULT_PROFILE],
        'lastUsedProfile': DEFAULT_PROFILE,
    }
    for k in GLOBAL_KEYS:
        if k in raw:
            migrated[k] = raw[k]
    return migrated

def save_settings_raw(data):
    """Write settings dict to disk."""
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)

def get_active_profile_data(settings):
    """Return the active profile's data dict, creating it if needed."""
    name = settings.get('activeProfile', DEFAULT_PROFILE)
    if 'profiles' not in settings:
        settings['profiles'] = {}
    if name not in settings['profiles']:
        settings['profiles'][name] = {}
    return settings['profiles'][name]


@app.route('/api/settings', methods=['GET'])
def api_settings_get():
    """Return current settings including active profile data."""
    try:
        raw      = load_settings_raw()
        settings = migrate_settings(raw)
        profile  = get_active_profile_data(settings)
        # Return flat view of active profile + globals for JS compatibility
        flat = {k: profile.get(k, {}) for k in PROFILE_KEYS}
        for k in GLOBAL_KEYS:
            flat[k] = settings.get(k)
        # Also return profile metadata for the UI
        flat['_profiles']      = list(settings.get('profileOrder', [DEFAULT_PROFILE]))
        flat['_activeProfile'] = settings.get('activeProfile', DEFAULT_PROFILE)
        flat['_lastUsed']      = settings.get('lastUsedProfile', DEFAULT_PROFILE)
        return jsonify({'ok': True, 'settings': flat})
    except Exception as e:
        return jsonify({'ok': True, 'settings': {}})


@app.route('/api/settings', methods=['POST'])
def api_settings_save():
    """Save active profile data and global settings."""
    try:
        body     = request.get_json(force=True)
        raw      = load_settings_raw()
        settings = migrate_settings(raw)
        profile  = get_active_profile_data(settings)
        # Save profile-specific keys
        for k in PROFILE_KEYS:
            if k in body:
                profile[k] = body[k]
        # Save global keys
        for k in GLOBAL_KEYS:
            if k in body:
                settings[k] = body[k]
        save_settings_raw(settings)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/profiles', methods=['GET'])
def api_profiles_get():
    """Return list of profiles and active profile name."""
    try:
        raw      = load_settings_raw()
        settings = migrate_settings(raw)
        order    = settings.get('profileOrder', [DEFAULT_PROFILE])
        return jsonify({
            'ok':           True,
            'profiles':     order,
            'active':       settings.get('activeProfile', DEFAULT_PROFILE),
            'lastUsed':     settings.get('lastUsedProfile', DEFAULT_PROFILE),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/profiles/switch', methods=['POST'])
def api_profiles_switch():
    """Switch to a different profile."""
    try:
        body     = request.get_json(force=True)
        name     = body.get('name', '').strip()
        raw      = load_settings_raw()
        settings = migrate_settings(raw)
        if name not in settings.get('profiles', {}):
            return jsonify({'ok': False, 'error': f'Profile "{name}" not found'}), 400
        settings['activeProfile']   = name
        settings['lastUsedProfile'] = name
        save_settings_raw(settings)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/profiles/create', methods=['POST'])
def api_profiles_create():
    """Create a new empty profile, or duplicate an existing one."""
    try:
        body          = request.get_json(force=True)
        name          = body.get('name', '').strip()
        duplicate_from = body.get('duplicateFrom', None)
        if not name:
            return jsonify({'ok': False, 'error': 'Profile name cannot be empty'}), 400
        if len(name) > 30:
            return jsonify({'ok': False, 'error': 'Profile name max 30 characters'}), 400
        raw      = load_settings_raw()
        settings = migrate_settings(raw)
        if name in settings.get('profiles', {}):
            return jsonify({'ok': False, 'error': f'Profile "{name}" already exists'}), 400
        if duplicate_from and duplicate_from in settings.get('profiles', {}):
            src = settings['profiles'][duplicate_from]
            import copy
            new_profile = copy.deepcopy(src)
        else:
            new_profile = {k: {} for k in PROFILE_KEYS}
        if 'profiles' not in settings:
            settings['profiles'] = {}
        settings['profiles'][name] = new_profile
        if 'profileOrder' not in settings:
            settings['profileOrder'] = [DEFAULT_PROFILE]
        settings['profileOrder'].append(name)
        save_settings_raw(settings)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/profiles/rename', methods=['POST'])
def api_profiles_rename():
    """Rename a profile."""
    try:
        body    = request.get_json(force=True)
        old     = body.get('old', '').strip()
        new     = body.get('new', '').strip()
        if not new:
            return jsonify({'ok': False, 'error': 'New name cannot be empty'}), 400
        if len(new) > 30:
            return jsonify({'ok': False, 'error': 'Profile name max 30 characters'}), 400
        raw      = load_settings_raw()
        settings = migrate_settings(raw)
        if old not in settings.get('profiles', {}):
            return jsonify({'ok': False, 'error': f'Profile "{old}" not found'}), 400
        if new in settings['profiles']:
            return jsonify({'ok': False, 'error': f'Profile "{new}" already exists'}), 400
        # Rename in profiles dict
        settings['profiles'][new] = settings['profiles'].pop(old)
        # Rename in order list
        order = settings.get('profileOrder', [])
        if old in order:
            order[order.index(old)] = new
        # Update active/lastUsed if needed
        if settings.get('activeProfile') == old:
            settings['activeProfile'] = new
        if settings.get('lastUsedProfile') == old:
            settings['lastUsedProfile'] = new
        save_settings_raw(settings)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/profiles/delete', methods=['POST'])
def api_profiles_delete():
    """Delete a profile. Cannot delete the last remaining profile."""
    try:
        body     = request.get_json(force=True)
        name     = body.get('name', '').strip()
        raw      = load_settings_raw()
        settings = migrate_settings(raw)
        profiles = settings.get('profiles', {})
        if name not in profiles:
            return jsonify({'ok': False, 'error': f'Profile "{name}" not found'}), 400
        if len(profiles) <= 1:
            return jsonify({'ok': False, 'error': 'Cannot delete the only profile'}), 400
        del profiles[name]
        order = settings.get('profileOrder', [])
        if name in order:
            order.remove(name)
        # If we deleted the active profile, switch to the first remaining
        if settings.get('activeProfile') == name:
            settings['activeProfile'] = order[0] if order else DEFAULT_PROFILE
        if settings.get('lastUsedProfile') == name:
            settings['lastUsedProfile'] = settings['activeProfile']
        save_settings_raw(settings)
        return jsonify({'ok': True, 'newActive': settings['activeProfile']})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/check_update', methods=['GET'])
def api_check_update():
    """Check GitHub for latest release version."""
    try:
        import urllib.request, urllib.error
        req = urllib.request.Request(GITHUB_API,
            headers={'User-Agent': f'EvonyGearOptimizer/{APP_VERSION}'})
        try:
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read())
        except urllib.error.HTTPError as e:
            if e.code == 404:
                # No releases published yet — treat as up to date
                return jsonify({'ok': True, 'current': APP_VERSION,
                                'latest': APP_VERSION, 'has_update': False,
                                'url': '', 'note': 'No releases found'})
            raise
        latest = data.get('tag_name', '').lstrip('v')
        url    = data.get('html_url', '')
        has_update = latest and latest != APP_VERSION
        return jsonify({'ok': True, 'current': APP_VERSION,
                        'latest': latest, 'has_update': has_update, 'url': url})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/sync_data', methods=['POST'])
def api_sync_data():
    """Download latest evony_data.xlsx from GitHub release asset and replace local copy."""
    try:
        import urllib.request, hashlib, shutil
        # Get release metadata to find asset URL
        api_req = urllib.request.Request(GITHUB_API,
            headers={'User-Agent': f'EvonyGearOptimizer/{APP_VERSION}',
                     'Accept': 'application/vnd.github+json'})
        with urllib.request.urlopen(api_req, timeout=8) as resp:
            release_data = json.loads(resp.read())
        asset_url = None
        for asset in release_data.get('assets', []):
            if asset['name'] == 'evony_data.xlsx':
                asset_url = asset['browser_download_url']
                break
        if not asset_url:
            return jsonify({'ok': False, 'error': 'Data asset not found on latest release'})
        # Download the asset
        data_req = urllib.request.Request(asset_url,
            headers={'User-Agent': f'EvonyGearOptimizer/{APP_VERSION}'})
        with urllib.request.urlopen(data_req, timeout=15) as resp:
            new_bytes = resp.read()
        with open(DATA_FILE, 'rb') as f:
            old_hash = hashlib.md5(f.read()).hexdigest()
        if hashlib.md5(new_bytes).hexdigest() == old_hash:
            return jsonify({'ok': True, 'updated': False})
        tmp = DATA_FILE + '.tmp'
        with open(tmp, 'wb') as f:
            f.write(new_bytes)
        shutil.move(tmp, DATA_FILE)
        invalidate_cache()
        return jsonify({'ok': True, 'updated': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/reload', methods=['POST'])
def api_reload():
    """Invalidate cache and force reload on next request."""
    invalidate_cache()
    return jsonify({'ok': True})


@app.route('/api/log', methods=['GET'])
def api_log():
    """Return contents of startup.log for troubleshooting."""
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                return jsonify({'ok': True, 'log': f.read()})
        return jsonify({'ok': True, 'log': '(log file not found)'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


def run_flask():
    """Run Flask in a background thread."""
    app.run(debug=False, port=5000, use_reloader=False)


def _startup_work(splash, webview):
    """
    Background thread: start Flask, warm cache, sync data, then
    open the main window and close the splash.
    All within the single webview event loop started by the splash.
    """
    import time, urllib.request, hashlib, shutil

    splash.set_status('Starting server...', 5)
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    time.sleep(0.6)

    splash.set_status('Loading gear data...', 30)
    try:
        get_cached_data()
    except Exception as e:
        splash.set_status(f'Warning: {e}', 30)
        time.sleep(1)

    splash.set_status('Checking for data updates...', 55)
    log(f'--- Startup data sync begin (v{APP_VERSION}) ---')
    log(f'DATA_FILE: {DATA_FILE}')
    log(f'GITHUB_API: {GITHUB_API}')
    try:
        # Step 1: get release metadata
        log('Step 1: fetching release metadata...')
        api_req = urllib.request.Request(
            GITHUB_API,
            headers={
                'User-Agent': f'EvonyGearOptimizer/{APP_VERSION}',
                'Accept': 'application/vnd.github+json',
            })
        with urllib.request.urlopen(api_req, timeout=8) as resp:
            release_data = json.loads(resp.read())
        log(f'Release tag: {release_data.get("tag_name")}')
        log(f'Assets: {[a["name"] for a in release_data.get("assets", [])]}')

        # Step 2: find asset URL
        asset_url = None
        for asset in release_data.get('assets', []):
            if asset['name'] == 'evony_data.xlsx':
                asset_url = asset['browser_download_url']
                break

        if not asset_url:
            log('ERROR: evony_data.xlsx not found in release assets')
            splash.set_status('Data check: no asset found', 65)
            time.sleep(0.5)
        else:
            log(f'Asset URL: {asset_url}')

            # Step 3: download
            log('Step 3: downloading asset...')
            data_req = urllib.request.Request(
                asset_url,
                headers={'User-Agent': f'EvonyGearOptimizer/{APP_VERSION}'})
            with urllib.request.urlopen(data_req, timeout=15) as resp:
                new_bytes = resp.read()
            log(f'Downloaded {len(new_bytes)} bytes')

            # Step 4: compare and replace
            with open(DATA_FILE, 'rb') as f:
                old_bytes = f.read()
            old_hash = hashlib.md5(old_bytes).hexdigest()
            new_hash = hashlib.md5(new_bytes).hexdigest()
            log(f'Old MD5: {old_hash}  New MD5: {new_hash}')

            if old_hash != new_hash:
                tmp = DATA_FILE + '.tmp'
                with open(tmp, 'wb') as f:
                    f.write(new_bytes)
                shutil.move(tmp, DATA_FILE)
                invalidate_cache()
                get_cached_data()
                log('Data file replaced successfully')
                splash.set_status('Data updated.', 65)
                time.sleep(0.3)
            else:
                log('Data file unchanged (hashes match)')
                splash.set_status('Data is current.', 65)

    except Exception as e:
        import traceback
        log(f'EXCEPTION: {e}')
        log(traceback.format_exc())
        splash.set_status('Data check skipped (offline?)', 65)
        time.sleep(0.2)
    log('--- Startup data sync end ---')

    splash.set_status('Opening interface...', 85)
    time.sleep(0.2)

    # Create the main window — still within the same webview event loop
    webview.create_window(
        title     = 'Evony Gear Optimizer v1.3.0',
        url       = 'http://127.0.0.1:5000',
        width     = 1400,
        height    = 860,
        min_size  = (900, 600),
        resizable = True,
    )
    splash.set_status('Ready!', 100)
    time.sleep(0.3)
    splash.close()   # destroy splash; main window keeps event loop alive


if __name__ == '__main__':
    import sys

    use_browser = '--browser' in sys.argv

    if use_browser:
        print("=" * 60)
        print("  Evony Gear Optimizer v1.3.0  [browser mode]")
        print(f"  Data file: {DATA_FILE}")
        print("  Opening browser at http://127.0.0.1:5000")
        print("=" * 60)
        threading.Timer(1.2, lambda: webbrowser.open('http://127.0.0.1:5000')).start()
        app.run(debug=False, port=5000)
    else:
        try:
            import webview
        except ImportError:
            # pywebview not available — fall back to browser
            threading.Timer(1.2, lambda: webbrowser.open('http://127.0.0.1:5000')).start()
            app.run(debug=False, port=5000)
        else:
            try:
                from splash import SplashScreen
                splash = SplashScreen()
                # _startup_work runs in background thread;
                # splash.run_in_thread starts webview event loop (blocks until all windows closed)
                t = threading.Thread(
                    target=_startup_work, args=(splash, webview), daemon=True)
                t.start()
                splash._open_window()   # starts webview event loop — blocks here
            except Exception as e:
                # Splash failed — open main window directly
                flask_thread = threading.Thread(target=run_flask, daemon=True)
                flask_thread.start()
                import time; time.sleep(1)
                webview.create_window(
                    title='Evony Gear Optimizer v1.3.0',
                    url='http://127.0.0.1:5000',
                    width=1400, height=860,
                    min_size=(900,600), resizable=True)
                webview.start()
